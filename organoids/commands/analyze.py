import click
import exif
import json
import os
import shapely
import tqdm
from PIL import Image  # For pixel data

from ..utils import end, start, status

@click.group()
def _analyze():
    pass

@_analyze.command()
@click.argument("directory", type=click.Path(exists=True), nargs=-1)
@click.option("--ext", default=".json", help="File extension to search for (default: .json)")
@click.option("--exif-ext", default=".jpg", help="File extension to extract EXIF from (default: .jpg)")
@click.option("--output", type=click.Path(), help="Output CSV file to write results (default: None, does not write to disk)")
def analyze(directory, ext, exif_ext, output):
    start("Scanning for files")
    todo = list(directory)
    found = []
    while todo:
        dir = todo.pop()
        for entry in os.listdir(dir):
            entry_path = os.path.join(dir, entry)
            if os.path.isdir(entry_path):
                todo.append(entry_path)
            elif entry.endswith(ext):
                found.append(entry_path)
    status(len(found), end='')
    end()
    
    data = {}
    for entry in tqdm.tqdm(found, desc="Parsing JSON and checking for shapes data"):
        with open(entry, 'rt') as f:
            d = json.load(f)
            if "shapes" in d:
                data[entry] = d
            else:
                print(f"Warning: {entry} has no shapes data")
    status(len(data), end='')
    end()
    
    start("Computing areas")
    measurements = {}
    for entry, d in tqdm.tqdm(data.items(), desc="Computing areas"):
        if output:
            day = os.path.basename(entry).split("d")[-1].split("_")[0]
        image_path = os.path.join(os.path.dirname(entry), d["imagePath"])
        
        pix_size = None
        mag = None
        exif_found = False
        
        # Check if the image has EXIF data
        if image_path.endswith(exif_ext):
            e = exif.Image(open(image_path, 'rb'))
            if e.has_exif and hasattr(e, "user_comment"):
                exif_found = True
                user_comment = json.loads(e.user_comment)
                pix_size = user_comment["effectivePixelSize"] / 10**6
                mag = user_comment["objectiveMag"]
            else:
                print(f"Warning: {image_path} referenced from {entry} has no valid EXIF data")

        # If no valid EXIF data, use pixel-based fallback
        if not exif_found:
            with Image.open(image_path) as img:
                width, height = img.size  # Get image dimensions
                pixel_area = width * height  # Calculate total pixel area
                pix_size = 1  # Assume each pixel has a size of 1 unit if no EXIF data
                mag = 1  # Default to no magnification if not specified
                print(f"Using pixel dimensions for {image_path}: {width}x{height} pixels")

        # Compute areas and update each shape's label
        for s in d["shapes"]:
            poly = shapely.geometry.Polygon(s["points"])
            area = (poly.area * pix_size) / mag
            
            # Update shape label to include area in the desired units
            area_label = f"Area: {area:.2f} mm²" if exif_found else f"Area: {area:.2f} pixels²"
            s["label"] = f"{s['label']} ({area_label})"  # Append area to the label
            s["area"] = area
            if output:
                measurements.setdefault(day, []).append(repr(area))
    end()
    start("Writing areas to disk")
    for entry, d in tqdm.tqdm(data.items(), desc="Writing areas to disk"):
        with open(entry, 'wt') as f:
            json.dump(d, f, indent=2)
    if output:
        days = sorted(int(day) for day in measurements.keys())
        with open(f"{output}.csv", 'w') as f:
            for day in days:
                f.write(f"{day},{','.join(measurements[day])}\n")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = directory[0]
        ws.cell(row=1, column=1, value="Day")
        max_len = max(len(meas) for meas in data.values())
        for i in range(max_len):
            ws.cell(row=1, column=i + 2, value=f"Measurement {i+1}")
        for row_idx, (day, measurements) in enumerate(data.items(), start=2):
            ws.cell(row=row_idx, column=1, value=day)
            for col_offset, value in enumerate(measurements):
                ws.cell(row=row_idx, column=col_offset + 2, value=value)
        wb.save(f"{output}.xlsx")

    end()
